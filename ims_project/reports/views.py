from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Avg, Count, Q
from accounts.decorators import not_student
from students.models import Student, Programme, Batch
from internships.models import InternshipRecord, MentorAssignment
from organisations.models import Organisation
from assessments.models import AssessmentMark
from assessments.calculations import calculate_student_score
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

@login_required
@not_student
def report_home(request):
    return render(request, 'reports/report_home.html')

@login_required
@not_student
def student_report(request):
    students = Student.objects.select_related('programme','batch').all()
    programme_id = request.GET.get('programme')
    batch_id = request.GET.get('batch')
    status = request.GET.get('status')
    if programme_id:
        students = students.filter(programme_id=programme_id)
    if batch_id:
        students = students.filter(batch_id=batch_id)
    if status:
        students = students.filter(status=status)
    programmes = Programme.objects.all()
    batches = Batch.objects.all()
    return render(request, 'reports/student_report.html', {
        'students': students, 'programmes': programmes, 'batches': batches,
    })

@login_required
@not_student
def marks_report(request):
    batch_id = request.GET.get('batch')
    programme_id = request.GET.get('programme')
    students = Student.objects.select_related('programme','batch').all()
    if batch_id:
        students = students.filter(batch_id=batch_id)
    if programme_id:
        students = students.filter(programme_id=programme_id)

    data = []
    for student in students:
        result = calculate_student_score(student)
        marks = [row['mark'] for row in result['regular_marks']]
        # Pad/truncate to exactly 8 columns for the table
        marks = (marks + [None] * 8)[:8]
        data.append({
            'student': student,
            'marks': marks,
            'regular_avg': result['regular_avg'],
            'assessment_mark': result['assessment_mark'],
            'final_score': result['final_score'],
            'is_provisional': result['is_provisional'],
        })

    programmes = Programme.objects.all()
    batches = Batch.objects.all()
    return render(request, 'reports/marks_report.html', {
        'data': data, 'programmes': programmes, 'batches': batches,
        'range8': range(1, 9),
    })

@login_required
@not_student
def org_report(request):
    orgs = Organisation.objects.annotate(intern_count=Count('internship_records')).order_by('-intern_count')
    return render(request, 'reports/org_report.html', {'orgs': orgs})

@login_required
@not_student
def mentor_report(request):
    from accounts.models import User
    mentors = User.objects.filter(role='faculty_mentor')
    data = []
    for mentor in mentors:
        assignments = MentorAssignment.objects.filter(faculty=mentor).select_related('student')
        pending = InternshipRecord.objects.filter(
            student__in=assignments.values('student'), verification_status='submitted'
        ).count()
        data.append({'mentor': mentor, 'assignments': assignments, 'pending': pending, 'count': assignments.values('student').distinct().count()})
    return render(request, 'reports/mentor_report.html', {'data': data})

@login_required
@not_student
def pending_report(request):
    pending_marks = InternshipRecord.objects.filter(
        verification_status='verified',
        assessment_mark__isnull=True
    ).select_related('student','organisation')
    pending_docs = InternshipRecord.objects.filter(
        Q(certificate='')|Q(certificate__isnull=True)
    ).select_related('student').filter(completion_status='completed')[:50]
    return render(request, 'reports/pending_report.html', {
        'pending_marks': pending_marks, 'pending_docs': pending_docs,
    })

@login_required
@not_student
def export_marks_excel(request):
    batch_id = request.GET.get('batch')
    programme_id = request.GET.get('programme')
    students = Student.objects.select_related('programme','batch').all()
    if batch_id:
        students = students.filter(batch_id=batch_id)
    if programme_id:
        students = students.filter(programme_id=programme_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Consolidation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1e3a5f")

    headers = (['Register No', 'Name', 'Programme', 'Batch']
               + [f'Intern {i}' for i in range(1, 9)]
               + ['Best 7 Avg (70%)', 'Assessment Mark (30%)', 'Final Score', 'Provisional?', 'Status'])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_num, student in enumerate(students, 2):
        result = calculate_student_score(student)
        ws.cell(row=row_num, column=1, value=student.register_number)
        ws.cell(row=row_num, column=2, value=student.name)
        ws.cell(row=row_num, column=3, value=student.programme.name)
        ws.cell(row=row_num, column=4, value=student.batch.name)

        marks = [row['mark'] for row in result['regular_marks']]
        marks = (marks + [None] * 8)[:8]
        for i, mark in enumerate(marks):
            ws.cell(row=row_num, column=5 + i, value=mark)

        ws.cell(row=row_num, column=13, value=result['regular_avg'])
        ws.cell(row=row_num, column=14, value=result['assessment_mark'])
        ws.cell(row=row_num, column=15, value=result['final_score'])
        ws.cell(row=row_num, column=16, value='Yes' if result['is_provisional'] else 'No')
        ws.cell(row=row_num, column=17, value=student.get_status_display())

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    # Add a note row explaining the formula
    note_row = students.count() + 3
    ws.cell(row=note_row, column=1,
            value="Final Score = 70% x (Best 7 of 8 regular internship average) + 30% x (assessment internship viva mark)")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="marks_report.xlsx"'
    wb.save(response)
    return response
