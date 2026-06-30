from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from .models import Student, BreakRecord, Programme, Batch
from .forms import StudentForm, BreakRecordForm, ProgrammeForm, BatchForm
from accounts.decorators import admin_required, not_student
from accounts.models import ActivityLog

@login_required
@not_student
def student_list(request):
    q = request.GET.get('q','')
    status = request.GET.get('status','')
    programme_id = request.GET.get('programme','')
    batch_id = request.GET.get('batch','')
    students = Student.objects.select_related('programme','batch').all()
    if q:
        students = students.filter(Q(register_number__icontains=q)|Q(name__icontains=q)|Q(email__icontains=q))
    if status:
        students = students.filter(status=status)
    if programme_id:
        students = students.filter(programme_id=programme_id)
    if batch_id:
        students = students.filter(batch_id=batch_id)
    programmes = Programme.objects.filter(is_active=True)
    batches = Batch.objects.filter(is_active=True)
    return render(request, 'students/student_list.html', {
        'students': students, 'q': q, 'status': status,
        'programmes': programmes, 'batches': batches,
        'statuses': Student.STATUS_CHOICES,
    })

@login_required
def student_detail(request, pk):
    if request.user.is_student_role:
        try:
            student = request.user.student_profile
            if student.pk != pk:
                messages.error(request, 'Access denied.')
                return redirect('dashboard')
        except Exception:
            messages.error(request, 'Student profile not found.')
            return redirect('dashboard')
    student = get_object_or_404(Student.objects.select_related('programme','batch'), pk=pk)
    internships = student.internship_records.select_related('organisation').order_by('internship_number')
    breaks = student.breaks.all()
    mentors = student.mentor_assignments.select_related('faculty').order_by('-effective_from')
    from assessments.models import InternshipMarks
    from django.db.models import Avg
    viva_avg = InternshipMarks.objects.filter(
        internship_record__student=student
    ).aggregate(avg=Avg('viva_marks'))['avg']
    return render(request, 'students/student_detail.html', {
        'student': student, 'internships': internships,
        'breaks': breaks, 'mentors': mentors, 'viva_avg': viva_avg,
    })

@login_required
@admin_required
def student_create(request):
    form = StudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        student = form.save()
        ActivityLog.objects.create(user=request.user, action='Created student', module='Students', record_id=str(student.pk), new_value=str(student))
        messages.success(request, f'Student {student.name} created successfully.')
        return redirect('student_list')
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Add Student'})

@login_required
@admin_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = StudentForm(request.POST or None, instance=student)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Student updated successfully.')
        return redirect('student_list')
    return render(request, 'students/student_form.html', {'form': form, 'title': f'Edit: {student.name}', 'student': student})

@login_required
@not_student
def break_create(request, student_pk):
    student = get_object_or_404(Student, pk=student_pk)
    form = BreakRecordForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        br = form.save(commit=False)
        br.student = student
        if br.start_date < student.degree_start_date or br.end_date > student.degree_end_date:
            messages.error(request, 'Break dates must fall within the degree period.')
        elif br.end_date <= br.start_date:
            messages.error(request, 'End date must be after start date.')
        else:
            br.save()
            messages.success(request, 'Break record added.')
            return redirect('student_detail', pk=student_pk)
    return render(request, 'students/break_form.html', {'form': form, 'student': student})

@login_required
@not_student
def programme_list(request):
    programmes = Programme.objects.all()
    return render(request, 'students/programme_list.html', {'programmes': programmes})

@login_required
@admin_required
def programme_create(request):
    form = ProgrammeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Programme created.')
        return redirect('programme_list')
    return render(request, 'students/programme_form.html', {'form': form, 'title': 'Add Programme'})

@login_required
@not_student
def batch_list(request):
    batches = Batch.objects.select_related('programme').all()
    return render(request, 'students/batch_list.html', {'batches': batches})

@login_required
@admin_required
def batch_create(request):
    form = BatchForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Batch created.')
        return redirect('batch_list')
    return render(request, 'students/batch_form.html', {'form': form, 'title': 'Add Batch'})


@login_required
@admin_required
def student_import(request):
    """
    Bulk student upload from Excel (.xlsx) or CSV, per SRS FR-ST-04.
    Expected columns (case-insensitive header match):
        register_number, name, email, mobile, programme_code, batch_name,
        degree_start_date, degree_end_date, status
    programme_code and batch_name must already exist (created via
    Students > Programmes / Batches first).
    """
    import openpyxl
    import csv
    import io
    from datetime import datetime

    results = None
    if request.method == 'POST' and request.FILES.get('import_file'):
        upload = request.FILES['import_file']
        rows = []
        errors_reading = []

        try:
            if upload.name.lower().endswith('.csv'):
                decoded = io.TextIOWrapper(upload.file, encoding='utf-8-sig')
                reader = csv.DictReader(decoded)
                rows = [{(k or '').strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]
            else:
                wb = openpyxl.load_workbook(upload, data_only=True)
                ws = wb.active
                header_row = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    rows.append({header_row[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(header_row))})
        except Exception as e:
            errors_reading.append(f'Could not read the file: {e}')

        created, skipped, row_errors = [], [], []

        def parse_date(value):
            if not value:
                return None
            if isinstance(value, str):
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                    try:
                        return datetime.strptime(value, fmt).date()
                    except ValueError:
                        continue
                return None
            try:
                return value.date()
            except AttributeError:
                return None

        for idx, row in enumerate(rows, start=2):
            reg_no = row.get('register_number', '')
            name = row.get('name', '')
            email = row.get('email', '')
            if not reg_no or not name:
                row_errors.append(f'Row {idx}: missing register_number or name — skipped.')
                continue
            if Student.objects.filter(register_number=reg_no).exists():
                skipped.append(f'Row {idx}: {reg_no} already exists — skipped.')
                continue

            programme = Programme.objects.filter(code__iexact=row.get('programme_code', '')).first()
            batch = Batch.objects.filter(name__iexact=row.get('batch_name', '')).first()
            if not programme or not batch:
                row_errors.append(f'Row {idx}: programme_code or batch_name not found — skipped.')
                continue

            start = parse_date(row.get('degree_start_date'))
            end = parse_date(row.get('degree_end_date'))
            if not start or not end:
                row_errors.append(f'Row {idx}: invalid or missing degree dates — skipped.')
                continue

            student = Student.objects.create(
                register_number=reg_no, name=name, email=email,
                mobile=row.get('mobile', ''), programme=programme, batch=batch,
                degree_start_date=start, degree_end_date=end,
                status=row.get('status', 'active') or 'active',
            )
            created.append(reg_no)

        if created:
            ActivityLog.objects.create(
                user=request.user, action=f'Bulk-imported {len(created)} students',
                module='Students', new_value=', '.join(created[:20])
            )

        results = {
            'created': created, 'skipped': skipped,
            'row_errors': row_errors + errors_reading,
        }
        if created:
            messages.success(request, f'{len(created)} student(s) imported successfully.')
        if not created and not row_errors and not skipped:
            messages.warning(request, 'No rows were found in the uploaded file.')

    return render(request, 'students/student_import.html', {'results': results})
